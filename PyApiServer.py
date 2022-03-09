import asyncio
import base64
import json
import os.path
import re
import time
from itertools import chain

import cv2
import imutils
import jiagu
import langid
import numpy as np
import paddleocr
import requests
from fastapi import FastAPI, status, Form
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from openpyxl import Workbook
from pypinyin import pinyin, Style
from selenium import webdriver
from sklearn.cluster import KMeans

import BpmDetector

# Do initialization here
api = FastAPI()
ocr = paddleocr.PaddleOCR(use_angle_cls=True, use_gpu=False)


# Api begin

@api.exception_handler(RequestValidationError)
async def doHandleValidationError(_request, _exc):
    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={'status': False, 'message': 'field required'}
    )


@api.post('/language')
async def doDetectLanguage(text: str = Form(...), encode: str = Form(...)):
    text = await getDecodedText(text, encode)
    return {'status': True, 'data': langid.classify(text)[0]}


@api.post('/pinyin')
async def doGetPinyin(text: str = Form(...), encode: str = Form(...)):
    text = await getDecodedText(text, encode)
    return {'status': True, 'data': ''.join([j[0] for j in pinyin(text.strip(), style=Style.FIRST_LETTER)])}


@api.post('/bpm')
async def getBPM(audio: str = Form(...)):
    if ':\\' in audio:
        audio = await convertWindowsPath(audio)
    loop = asyncio.get_event_loop()
    bpm = await loop.run_in_executor(None, BpmDetector.detectWav, audio)
    return {'status': True, 'data': bpm}


@api.post('/sentiment')
async def doDetectSentiment(text: str = Form(...)):
    return {'status': True, 'data': jiagu.sentiment(text)[0] == 'positive'}


@api.post('/hair_color')
async def getHairColor(image: str = Form(...)):
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, handleHairColor, image)
    return {'status': True, 'data': ';'.join([','.join(map(str, reversed(j))) for j in result])}


@api.post('/ocr')
async def getOCR(url: str = Form(...)):
    r, result = "", ocr.ocr(url, cls=True)
    if result is not None:
        for j in result:
            r += j[-1][0] + " "
    return {'status': True, 'data': r}


@api.get('/maimai_aliases')
async def getMaimaiAliases():
    wb = TencentDocDownload().fetch("https://docs.qq.com/sheet/DWGNNYUdTT01PY2N1")
    aliases = {}
    for row in wb.worksheets[2].iter_rows():
        if row[0].value == "歌名":
            continue
        if str(row[0].value).replace('\xa0', ' ') in aliases.keys():
            aliases[row[0].value.replace('\xa0', ' ')] = list(set(
                aliases[row[0].value.replace('\xa0', ' ')] +
                [str(cell.value) for cell in row[1:] if cell.value != ""]
            ))
        else:
            aliases[str(row[0].value)] = [str(cell.value) for cell in row[1:] if cell.value != ""]
    return {'status': True, 'data': json.dumps(aliases)}


@api.post('/is_lt')
async def isLt(image: str = Form(...)):
    return {'status': True, 'data': str(LTDetector().isLt(image))}


@api.post('/nsfw_detect')
async def nsfwDetect(image: str = Form(...)):
    loop = asyncio.get_event_loop()
    return {'status': True, 'data': await loop.run_in_executor(None, baiduNsfwDetect, image)}


# Api end


async def convertWindowsPath(path: str):
    parts = path.split(':\\')
    return '/mnt/' + parts[0].lower() + '/' + parts[1].replace('\\', '/')


async def getDecodedText(text: str, encode: str):
    encode = encode.lower()
    if encode in ['utf-8', 'utf8', 'unicode']:
        text = text.encode('utf-8').decode('unicode_escape')
    elif encode == 'base64':
        text = base64.b64decode(text).decode('utf-8')
    return text


# Based on https://www.pyimagesearch.com/2014/05/26/opencv-python-k-means-color-clustering/
def centroidHistogram(clt):
    numLabels = np.arange(0, len(np.unique(clt.labels_)) + 1)
    (hist, _) = np.histogram(clt.labels_, bins=numLabels)
    hist = hist.astype("float")
    hist /= hist.sum()
    return hist


def getMainColor(hist, centroids):
    return sorted(zip(hist, centroids), key=lambda x: -x[0])[0][1]


def handleHairColor(filename, cascade_file="./lbpcascade_animeface.xml"):
    if not os.path.isfile(cascade_file):
        raise RuntimeError("%s: not found" % cascade_file)

    cascade = cv2.CascadeClassifier(cascade_file)
    image = cv2.imread(filename, cv2.IMREAD_COLOR)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)

    faces = cascade.detectMultiScale(gray,
                                     # detector options
                                     scaleFactor=1.1,
                                     minNeighbors=1,
                                     minSize=(24, 24))
    if len(faces) == 0:
        faces = [(0, 0, image.shape[1], image.shape[0])]
    results = []
    for (x, y, w, h) in faces:
        now = image[y:y + h // 4, x:x + w]
        now = now.reshape((now.shape[0] * now.shape[1], 3))
        clt = KMeans(n_clusters=6)
        clt.fit(now)
        results.append(getMainColor(centroidHistogram(clt), clt.cluster_centers_))
    return results


# Based on https://github.com/kuloPo/TencentDocDownload
class TencentDocDownload:
    apiOpendoc = "https://docs.qq.com/dop-api/opendoc"
    apiSheet = "https://docs.qq.com/dop-api/get/sheet"

    def init(self, url):
        initUrl = url
        if initUrl.find('?') != -1:
            initUrl = initUrl[:initUrl.find('?')]

        initText = requests.get(initUrl).text

        t = re.search(r"&t=(\d+)\"", initText).group(1)
        docId = re.search(r"/sheet/(.+)\??", initUrl).group(1)

        opendocParams = {
            "id": docId,
            "normal": "1",
            "outformat": "1",
            "startrow": "0",
            "endrow": "60",
            "wb": "1",
            "nowb": "0",
            "callback": "clientVarsCallback",
            "xsrf": "",
            "t": t
        }
        opendocText = requests.get(self.apiOpendoc, params=opendocParams).text
        opendocJson = self.readCallback(opendocText)

        title = opendocJson["clientVars"]["title"]
        tabs = opendocJson["clientVars"]["collab_client_vars"]["header"][0]["d"]

        return title, tabs, opendocParams

    def readSheet(self, sheet, opendocParams):
        opendocParams["tab"] = sheet
        opendocText = requests.get(self.apiOpendoc, params=opendocParams).text
        opendocJson = self.readCallback(opendocText)
        max_row = opendocJson["clientVars"]["collab_client_vars"]["maxRow"]
        maxCol = opendocJson["clientVars"]["collab_client_vars"]["maxCol"]
        padId = opendocJson["clientVars"]["collab_client_vars"]["globalPadId"]
        rev = opendocJson["clientVars"]["collab_client_vars"]["rev"]

        sheetParams = {
            "tab": sheet,
            "padId": padId,
            "subId": sheet,
            "outformat": "1",
            "startrow": "0",
            "endrow": max_row,
            "normal": "1",
            "preview_token": "",
            "nowb": "1",
            "rev": rev
        }
        sheetText = requests.get(self.apiSheet, params=sheetParams).text
        sheetJson = json.loads(sheetText)
        sheetContent = {}
        for tmpClass in sheetJson["data"]["initialAttributedText"]["text"][0]:
            if type(tmpClass[0]) == dict and "c" in tmpClass[0].keys():
                if len(tmpClass[0]["c"]) > 1 and type(tmpClass[0]["c"][1]) == dict:
                    temp = tmpClass[0]["c"][1]  # type: dict
                    for k, v in temp.items():
                        if k.isdigit() and type(v) == dict:
                            sheetContent[k] = v
        return sheetContent, maxCol

    @staticmethod
    def readCallback(text):
        content = re.search(r"clientVarsCallback\(\"(.+)\"\)", text).group(1)
        content = content.replace("&#34;", "\"")
        content = content.replace(r'\\"', r"\\'")
        return json.loads(content)

    def fetch(self, url):
        title, tabs, opendocParams = self.init(url)
        wb = Workbook()
        for tab in tabs:
            tabId = tab["id"]
            name = tab["name"]
            sheetContent, maxCol = self.readSheet(tabId, opendocParams)
            row = []
            ws = wb.create_sheet(name)
            for k, v in sheetContent.items():
                if int(k) % maxCol == 0 and k != '0':
                    ws.append(row)
                    row = []
                if '2' in v:
                    row.append(v['2'][1])
                else:
                    row.append("")
        emptyWs = wb["Sheet"]
        wb.remove(emptyWs)
        return wb


class LTDetector:
    @staticmethod
    def pHash(img):
        if len(img.shape) == 3:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        img = cv2.resize(img, (20, 20), interpolation=cv2.INTER_CUBIC)
        h, w = img.shape[:2]
        vis0 = np.zeros((h, w), np.float32)
        vis0[:h, :w] = img
        vis1 = cv2.dct(cv2.dct(vis0))
        vis1.resize((20, 20), refcheck=False)
        img_list = list(chain.from_iterable(vis1))
        avg = sum(img_list) * 1. / len(img_list)
        avg_list = ['0' if i < avg else '1' for i in img_list]
        return ''.join(['%x' % int(''.join(avg_list[x:x + 4]), 2) for x in range(0, 20 * 20, 4)])

    @staticmethod
    def hammingDist(s1, s2):
        assert len(s1) == len(s2)
        return sum([ch1 != ch2 for ch1, ch2 in zip(s1, s2)])

    @staticmethod
    def enhance(img):
        alpha = 2.5
        beta = 0
        return cv2.convertScaleAbs(img, alpha=alpha, beta=beta)

    def imgSimilarity(self, a, b):
        return 1 - self.hammingDist(self.pHash(self.enhance(a)), self.pHash(self.enhance(b))) * 1. / (16 * 16 / 4)

    def isLt(self, path):
        templateRaw = cv2.imread("/share/bot/lt.jpg")
        template = cv2.cvtColor(templateRaw, cv2.COLOR_BGR2GRAY)
        template = cv2.Canny(template, 50, 200)
        (tH, tW) = template.shape[:2]
        image = cv2.imread(path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        found = None
        if self.grayPercent(image) > 25:
            return 0.0
        for scale in np.linspace(0.2, 1.0, 20)[::-1]:
            resized = imutils.resize(gray, width=int(gray.shape[1] * scale))
            r = gray.shape[1] / float(resized.shape[1])
            if resized.shape[0] < tH or resized.shape[1] < tW:
                break
            edged = cv2.Canny(resized, 50, 200)
            result = cv2.matchTemplate(edged, template, cv2.TM_CCOEFF)
            (_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
            if found is None or maxVal > found[0]:
                found = (maxVal, maxLoc, r)
        if found is not None:
            (_, maxLoc, r) = found
            (startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
            (endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
            if (abs(endX - startX - tW) > 10) and (abs(endY - startY - tH) > 10) and \
                    (endX - startX < 43 or endY - startY < 43):
                return 0.0
            # draw a bounding box around the detected result and display the image
            sub = image[startY:endY, startX:endX]
            if self.grayPercent(sub):
                return self.imgSimilarity(sub, templateRaw)
            else:
                return 0.0
        return 0.0

    @staticmethod
    def grayPercent(img):
        img2 = cv2.cvtColor(cv2.cvtColor(img, cv2.COLOR_RGB2GRAY), cv2.COLOR_GRAY2RGB)
        diff = 0
        for i in range(img.shape[0]):
            for j in range(img.shape[1]):
                for k in range(img.shape[2]):
                    diff += abs(int(img[i, j, k]) - int(img2[i, j, k])) ** 2
        return diff / (img.shape[0] * img.shape[1] * img.shape[2]) < 5


def baiduNsfwDetect(imageUrl):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    browser = webdriver.Chrome(options=chrome_options)
    browser.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument',
                            {'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'})
    browser.get('https://ai.baidu.com/tech/imagecensoring')
    time.sleep(1)
    browser.find_element_by_class_name('image-url').send_keys(imageUrl)
    browser.find_element_by_class_name('image-button-active').click()
    time.sleep(2)
    result = browser.find_element_by_class_name('demo-json-content').text
    browser.quit()
    return result


if __name__ == '__main__':
    import uvicorn

    uvicorn.run(app=api, host="0.0.0.0", port=10090)
